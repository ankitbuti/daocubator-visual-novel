#!/usr/bin/env python3
"""Generate the art brief (backgrounds + character sprites) as an .xlsx sheet.

Output: vault/06 Assets/Art Assets.xlsx  (tabs: Backgrounds, Characters, Sprites, Specs & Notes)
Data is grounded in the actual `scene bg X` and `show <char> <expr>` refs in game/*.rpy.
Regenerate any time: python3 scripts/gen_asset_sheet.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = os.path.join(os.path.dirname(__file__), "..", "vault", "06 Assets", "Art Assets.xlsx")

# ---------------------------------------------------------------- data

# filename, uses, where it appears, what to depict, status
BACKGROUNDS = [
    ("bg conference.png", 13, "Alex's intro, Vera's mixer, Atlas's 500 ETH offer + treasury debate, Proposal #17, final vote, whale negotiation, acquisition offer",
     "Crypto-conference hall / pitch stage — big screen with DAO/ETH banners, seating. The public, formal, high-stakes room.", "placeholder"),
    ("bg discord.png", 8, "Jordan's intro, community launch, the leak, delegation debate, wallet-cluster warning, community defense, Death-by-Democracy ending",
     "Stylized Discord server at 2-3am — channel/member rail, chat column, dark glow. The online / community space.", "placeholder"),
    ("bg coworking.png", 7, "Opening / venture choice, Maya's intro + equity talk, Maya's 3x crisis, morning-after triage, Sustainable ending",
     "Startup co-working office — desks, monitors, whiteboard, window. Where the team actually works.", "placeholder"),
    ("bg blockchain.png", 7, "Cold open, TGE / mint night, token design, attack intro, 51% / spiral vectors, Hostile-Takeover ending",
     "Abstract on-chain visualization — glowing node graph + edges, cyan/magenta on black. 'The chain.'", "placeholder"),
    ("bg apartment.png", 7, "Spectre's 2:47am DM, audit ask, 4am burnout mirror, phish setup, Jordan's 3am call, forced-rest",
     "Late-night apartment — laptop glow, city lights through a window. The alone / personal / exhausted space.", "placeholder"),
    ("bg factionmap.png", 4, "Ch3 faction map, the 76.2% pie-chart moment, ending recap",
     "Governance dashboard / diagram — faction circles (Decentralists / Pragmatists / Opportunists), pie charts. More UI than place.", "placeholder"),
    ("bg rooftop.png", 2, "Vera's closing argument, Golden 'Collective Prosperity' ending",
     "Rooftop at dawn — city skyline, magenta->orange sky. The resolution / hope beat.", "placeholder"),
    ("bg signing.png", 1, "Ch4a 'verify the hex' set piece",
     "A hardware-wallet / MetaMask signing modal — 'SIGN TRANSACTION?', summary rows, an [EXPAND RAW DATA] hex block, green Confirm / red Reject. A UI mockup, not a location.", "placeholder"),
    ("(scene black)", "-", "Burnout-Crash ending + a few transitions",
     "No asset needed — Ren'Py generates solid black automatically. Listed for completeness.", "built-in"),
]

# character, role, archetype, accent hex, base look, design notes
CHARACTERS = [
    ("Maya", "Technical co-founder", "The Builder", "#55FFFF",
     "Practical; laptop always open, glasses, dry-but-kind. Reads as competent and a little tired.",
     "Load-bearing character — appears most. Cyan accent."),
    ("Jordan", "Community lead", "The True Believer", "#FF55FF",
     "Warm, expressive, DAO merch, meme energy; the heart of the room.",
     "Idealist with 2:47am energy. Magenta accent."),
    ("Alex", "Treasury / finance", "The Mercenary", "#FFAA55",
     "Sharp suit, always checking prices, guarded charm.",
     "Ethically flexible; loyalty stays ambiguous until Ch4. Orange accent."),
    ("Vera", "Mentor (ex-union organizer / angel)", "The Organizer", "#AA00AA",
     "Older, wry, seen-it-all; calm authority.",
     "Historical counterpoint to every theme. Purple accent."),
    ("Spectre", "Anonymous security contributor", "The Ghost", "#00AAAA",
     "Hooded, face in shadow, known only by PFP; screen-glow lighting.",
     "HARD RULE: stays hooded in EVERY expression — never a full face, in any scene or ending. Cyan accent."),
    ("Atlas", "Benefactor-antagonist (whale)", "The Whale", "#FFFF55",
     "Luxurious, mysterious wealth, unhurried power.",
     "Never a cartoon villain — poised, right about something in every scene. Yellow accent."),
]

# character, expression, emotion/pose note, in_script ('yes' referenced now / 'base' planned)
SPRITES = [
    ("Maya", "neutral", "Calm, attentive.", "yes"),
    ("Maya", "focused", "Heads-down, problem-solving.", "yes"),
    ("Maya", "tired", "Burned out, running on fumes.", "yes"),
    ("Maya", "proud", "Quiet satisfaction.", "yes"),
    ("Maya", "alarmed", "Red-alert; something's wrong.", "yes"),
    ("Jordan", "neutral", "Baseline warm.", "base"),
    ("Jordan", "hyped", "Thrilled, arms-up energy.", "yes"),
    ("Jordan", "hurt", "Let down; taking it personally.", "yes"),
    ("Jordan", "determined", "Resolute, rallying the room.", "yes"),
    ("Jordan", "exhausted", "3am, drained.", "yes"),
    ("Alex", "neutral", "Cool, appraising.", "yes"),
    ("Alex", "smirk", "Sly, amused.", "yes"),
    ("Alex", "calculating", "Doing the math on you.", "yes"),
    ("Alex", "sincere", "A rare genuine moment.", "yes"),
    ("Alex", "defensive", "Caught out, guarded.", "base"),
    ("Vera", "neutral", "Steady.", "base"),
    ("Vera", "wry", "Dry half-smile.", "yes"),
    ("Vera", "stern", "Warning you.", "yes"),
    ("Vera", "warm", "Approving; proud of you.", "yes"),
    ("Spectre", "neutral", "Hooded, still.", "yes"),
    ("Spectre", "typing", "Screen-glow, mid-message.", "yes"),
    ("Spectre", "alert", "Something detected.", "yes"),
    ("Spectre", "vulnerable", "A crack in the armor — still hooded.", "yes"),
    ("Atlas", "neutral", "Poised.", "base"),
    ("Atlas", "generous", "Magnanimous public face.", "yes"),
    ("Atlas", "guarded", "Calculating, in private.", "yes"),
    ("Atlas", "predatory", "Mask off — calm menace.", "yes"),
]

SPECS = [
    ("Backgrounds", "1920 x 1080. Files named 'bg <name>.png' in game/images/ (the space matters)."),
    ("Sprites", "~2/3-body character art, transparent PNG, roughly 560 x 880+ scaled onto the 1920x1080 stage; left/center/right slots. Files named '<character> <expression>.png' (lowercase, space between)."),
    ("Naming rule", "Ren'Py auto-image: file 'maya focused.png' is shown in script as 'show maya focused'. Keep exact names or the script breaks."),
    ("Style", "PC-98 16-color palette (deep blue / cyan / magenta / orange / cream on black), ordered dithering, sharp 1px edges, dramatic 90s anime shading."),
    ("Accent colors", "Each character owns one palette accent (see Characters tab): Maya cyan, Jordan magenta, Alex orange, Vera purple, Spectre cyan, Atlas yellow."),
    ("Priority", "'yes' in the Sprites tab = referenced by the script right now (make these first: 23). 'base' = neutral/defensive base the design calls for but the current script doesn't show yet (4). Total 27."),
    ("Location", "Drop finished files in game/images/, replacing the current scripted placeholders (same filenames)."),
    ("Palette source", "vault/06 Assets/Art Style Guide.md"),
    ("The Player", "No sprite — first-person POV, never shown on screen."),
]

# ---------------------------------------------------------------- styling

HEADER_FILL = PatternFill("solid", fgColor="1A1A4E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1A1A4E")
MONO = Font(name="Menlo", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def table(ws, title, subtitle, headers, rows, widths, wrap_cols, mono_col=0, hrow=4):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="555555")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
    for i, row in enumerate(rows):
        r = hrow + 1 + i
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP if c in wrap_cols else TOP
            cell.border = BORDER
            if c == mono_col:
                cell.font = MONO
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{hrow + 1}"


def main():
    wb = Workbook()

    ws = wb.active
    ws.title = "Backgrounds"
    table(ws, 'DAOCUBATOR / "Get Rich Together" — Background Assets',
          "8 backgrounds (grounded in the game's scene bg calls). 1920x1080, PC-98 palette. See Specs tab.",
          ["Filename", "Uses", "Where it appears", "What to depict", "Status"],
          BACKGROUNDS, {"A": 20, "B": 7, "C": 46, "D": 60, "E": 13}, wrap_cols={3, 4}, mono_col=1)

    wsc = wb.create_sheet("Characters")
    table(wsc, "Character design reference",
          "6 characters (the Player has no sprite). One accent color each; see Sprites tab for the per-expression file list.",
          ["Character", "Role", "Archetype", "Accent", "Base look", "Design notes"],
          CHARACTERS, {"A": 12, "B": 26, "C": 18, "D": 10, "E": 46, "F": 40}, wrap_cols={5, 6})

    wss = wb.create_sheet("Sprites")
    sprite_rows = [(f"{ch.lower()} {ex}.png", ch, ex, note, ins) for (ch, ex, note, ins) in SPRITES]
    table(wss, "Character sprites to create (27)",
          "One row = one PNG. 'yes' = used by the script now (23, do first). 'base' = neutral/defensive base the design calls for (4).",
          ["Filename", "Character", "Expression", "What it shows", "In script?", "Status"],
          [(fn, ch, ex, note, ins, "placeholder") for (fn, ch, ex, note, ins) in sprite_rows],
          {"A": 22, "B": 11, "C": 13, "D": 40, "E": 11, "F": 13}, wrap_cols={4}, mono_col=1)

    wsp = wb.create_sheet("Specs & Notes")
    table(wsp, "Technical specs & conventions", "",
          ["Item", "Detail"], SPECS, {"A": 16, "B": 96}, wrap_cols={2}, hrow=3)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("wrote", os.path.abspath(OUT))
    print(f"backgrounds: {len([b for b in BACKGROUNDS if b[0].startswith('bg')])}, "
          f"characters: {len(CHARACTERS)}, sprites: {len(SPRITES)} "
          f"({sum(1 for s in SPRITES if s[3]=='yes')} in-script + {sum(1 for s in SPRITES if s[3]=='base')} base)")


if __name__ == "__main__":
    main()
